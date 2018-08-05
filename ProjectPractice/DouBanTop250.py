import re
import requests
from bs4 import BeautifulSoup
import openpyxl


def open_url(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.84 Safari/537.36'
    }
    res = requests.get(url, headers=headers)
    return res


def find_movies(res):
    soup = BeautifulSoup(res.text, 'html.parser')
    movies = []
    targets = soup.find_all("div", class_="hd")
    for each in targets:
        movies.append(each.a.span.text)

    ranks = []
    targets = soup.find_all("span", class_="rating_num")
    for each in targets:
        ranks.append(each.text)

    messages = []
    targets = soup.find_all("div", class_="bd")
    for each in targets:
        try:
            messages.append(each.p.text.split('\n')[1].strip() + each.p.text.split('\n')[2].strip())
        except:
            continue

    result = []
    length = len(movies)
    for i in range(length):
        result.append([movies[i], ranks[i], messages[i]])
        # result.append(movies[i] + ranks[i] + messages[i] + '\n')

    return result


def find_depth(res):
    soup = BeautifulSoup(res.text, 'html.parser')
    depth = soup.find('span', class_='next').previous_sibling.previous_sibling.text
    return int(depth)


def save_to_excel(result):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "电影名称"
    ws['B1'] = "评分"
    ws['C1'] = "资料"

    for each in result:
        ws.append(each)

    wb.save("豆瓣TOP电影.xlsx")


def main():
    host = "https://movie.douban.com/top250"
    res = open_url(host)
    depth = find_depth(res)

    result = []
    for i in range(depth):
        url = host + '/?start=' + str(25 * i)
        res = open_url(url)
        result.extend(find_movies(res))

    save_to_excel(result)
    '''
    with open("豆瓣TOP250.txt", 'w', encoding='utf-8') as f:
        for each in result:
            f.write(each)
    '''


if __name__ == '__main__':
    main()
